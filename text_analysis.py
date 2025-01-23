import os
import pandas as pd
import nltk
from nltk.sentiment.vader import SentimentIntensityAnalyzer
from collections import Counter
import re
import spacy
from textblob import TextBlob
import subprocess
import sys
from openpyxl import load_workbook
import matplotlib.pyplot as plt

# Sprawdzanie, czy zasób VADER jest już pobrany
try:
    nltk.data.find('vader_lexicon')
except LookupError:
    nltk.download('vader_lexicon')

# Sprawdzanie, czy model en_core_web_sm jest zainstalowany
try:
    spacy.load("en_core_web_sm")
except OSError:
    print("Model en_core_web_sm nie jest zainstalowany. Instaluję...")
    subprocess.check_call([sys.executable, "-m", "spacy", "download", "en_core_web_sm"])

# Ładowanie modelu spaCy
nlp = spacy.load("en_core_web_sm")

# Funkcja do zapisywania wyników analizy do pliku CSV
def save_to_csv(song_name, analysis_results, analysis_type):
    file_name = "analiza_piosenki.xlsx"

    # Debug: Wydrukowanie ścieżki, gdzie zapisujemy plik
    print(f"Zapisuję do pliku: {os.path.abspath(file_name)}")

    # Przygotowanie tytułu kolumny w formacie "Analiza: {wybrana_analiza} - Piosenka {nazwa_piosenki}"
    column_name = f"Analiza: {analysis_type} - Piosenka {song_name}"

    # Jeśli wynik analizy to słownik (np. wynik sentymentu)
    if isinstance(analysis_results, dict):
        # Tworzymy osobne kolumny dla każdego klucza i wartości
        keys_column_name = f"{column_name} - Klucz"
        values_column_name = f"{column_name} - Wartość"

        # Przekształcamy słownik na dwie kolumny (klucz i wartość)
        analysis_results = pd.DataFrame(list(analysis_results.items()), columns=[keys_column_name, values_column_name])
        print(f"Przekształcone wyniki analizy (słownik):\n{analysis_results}")
    # Jeśli wynik analizy to lista, zapisujemy ją w formie pojedynczych wierszy
    elif isinstance(analysis_results, list):
        analysis_results = pd.DataFrame({column_name: analysis_results})
        print(f"Przekształcone wyniki analizy (lista):\n{analysis_results}")
    # Jeśli wynik analizy to krotka (np. analiza składniowa lub nazw własnych)
    elif isinstance(analysis_results, tuple):
        # Rozdzielamy krotkę na oddzielne kolumny
        # Każdy element krotki będzie zapisywany w oddzielnej kolumnie
        analysis_results = pd.DataFrame([analysis_results], columns=[f"{column_name} - Kolumna {i+1}" for i in range(len(analysis_results))])
        print(f"Przekształcone wyniki analizy (krotka):\n{analysis_results}")
    # Jeśli wynik analizy to pojedynczy tekst (np. błędy lub proste komunikaty)
    elif isinstance(analysis_results, str):
        # Przekształcamy tekst na listę, by można było go zapisać do CSV
        analysis_results = pd.DataFrame({column_name: [analysis_results]})
        print(f"Przekształcone wyniki analizy (tekst):\n{analysis_results}")

    # Zapisz do istniejącego pliku Excel lub stwórz nowy
    if os.path.exists(file_name):
        try:
            # Załaduj istniejący plik Excel
            with pd.ExcelWriter(file_name, engine='openpyxl', mode='a') as writer:
                # Sprawdź, czy arkusz już istnieje
                book = load_workbook(file_name)
                if analysis_type in book.sheetnames:
                    # Jeśli arkusz już istnieje, nadpisz go nowymi danymi
                    print(f"Arkusz '{analysis_type}' już istnieje, nadpisuję go.")
                # Zapisz dane do nowego arkusza w istniejącym pliku
                analysis_results.to_excel(writer, sheet_name=analysis_type, index=False)
                print(f"Wyniki zapisane w arkuszu '{analysis_type}' w pliku {file_name}")
        except Exception as e:
            print(f"Błąd przy zapisywaniu do istniejącego pliku: {e}")
    else:
        # Zapisz dane do nowego pliku Excel
        with pd.ExcelWriter(file_name, engine='openpyxl') as writer:
            analysis_results.to_excel(writer, sheet_name=analysis_type, index=False)
            print(f"Wyniki zapisane w nowym pliku {file_name}")


# Funkcja do rozdzielania wyników analizy nazw własnych na kolumny
def split_named_entity_recognition(entities):
    # Tworzymy listę, która będzie przechowywać rozdzielone wyniki
    split_results = []

    # Dla każdego rozpoznanego bytu w wynikach analizy
    for entity, label in entities:
        split_results.append({
            'Nazwa własna': entity,
            'Typ bytu': label
        })

    # Przekształcamy listę wyników na DataFrame
    df = pd.DataFrame(split_results)
    return df


def split_syntactic_analysis(analysis_results):
    # Tworzymy listę, która będzie przechowywać rozdzielone wyniki
    split_results = []

    # Dla każdego elementu krotki w wynikach analizy
    for result in analysis_results:
        # Rozdzielamy krotkę na poszczególne elementy: słowo, część mowy, lemat, zależność i główne słowo
        word, pos, lemma, dep, head = result
        split_results.append({
            'Słowo': word,
            'Część mowy': pos,
            'Lemat': lemma,
            'Zależność składniowa': dep,
            'Główne słowo': head
        })

    # Przekształcamy listę wyników na DataFrame
    df = pd.DataFrame(split_results)
    return df

# Funkcja 1: Analiza sentymentu
def sentiment_analysis(song):
    sia = SentimentIntensityAnalyzer()
    sentiment_score = sia.polarity_scores(song)
    return sentiment_score


# Funkcja 2: Analiza częstotliwości występowania słów
def word_frequency(song):
    song = song.lower()
    song = re.sub(r'[^\w\s]', '', song)
    doc = nlp(song)
    word_counts = Counter([token.text for token in doc if token.is_alpha])
    return word_counts


# Funkcja 3: Analiza n-gramów
def generate_ngrams(song, n=3):
    song = song.replace("\n", " ")
    doc = nlp(song)
    ngrams = []
    for i in range(len(doc) - n + 1):
        ngram = ' '.join([doc[i + j].text for j in range(n)])
        ngrams.append(ngram)
    ngram_counts = Counter(ngrams)
    return ngram_counts

# Funkcja 4: Analiza rozpoznawania nazw własnych
def named_entity_recognition(song):
    doc = nlp(song)
    entities = [(entity.text, entity.label_) for entity in doc.ents]

    # Rozdzielamy wyniki analizy nazw własnych na kolumny
    return split_named_entity_recognition(entities)  # Zwrócenie DataFrame

# Funkcja 5: Analiza sentymentu z użyciem TextBlob
def textblob_sentiment_analysis(song):
    blob = TextBlob(song)
    polarity = blob.sentiment.polarity
    subjectivity = blob.sentiment.subjectivity
    return polarity, subjectivity


# Funkcja 6: Analiza składniowa
def syntactic_analysis(song):
    doc = nlp(song)
    analysis = []
    for token in doc:
        analysis.append((token.text, token.pos_, token.lemma_, token.dep_, token.head.text))

    # Rozdzielamy wyniki analizy składniowej na kolumny
    return split_syntactic_analysis(analysis)  # Zwrócenie DataFrame

# Funkcja 2: Analiza częstotliwości występowania słów
def word_frequency(song):
    song = song.lower()
    song = re.sub(r'[^\w\s]', '', song)
    doc = nlp(song)
    word_counts = Counter([token.text for token in doc if token.is_alpha])
    return word_counts

# Funkcja do tworzenia wykresu z top 10 najczęstszych słów
def plot_word_frequency(word_counts):
    top_10 = word_counts.most_common(10)
    words, counts = zip(*top_10)
    plt.figure(figsize=(10, 6))
    plt.bar(words, counts)
    plt.title('Top 10 Najczęstszych Słów')
    plt.xlabel('Słowa')
    plt.ylabel('Częstotliwość')
    plt.xticks(rotation=45)
    plt.show()

# Funkcja 3: Analiza n-gramów
def generate_ngrams(song, n=3):
    song = song.replace("\n", " ")
    doc = nlp(song)
    ngrams = []
    for i in range(len(doc) - n + 1):
        ngram = ' '.join([doc[i + j].text for j in range(n)])
        ngrams.append(ngram)
    ngram_counts = Counter(ngrams)
    return ngram_counts

# Funkcja do tworzenia wykresu z top 10 najczęstszych n-gramów
def plot_ngrams(ngram_counts):
    top_10 = ngram_counts.most_common(10)
    ngrams, counts = zip(*top_10)
    plt.figure(figsize=(10, 6))
    plt.bar(ngrams, counts)
    plt.title('Top 10 Najczęstszych N-gramów')
    plt.xlabel('N-gramy')
    plt.ylabel('Częstotliwość')
    plt.xticks(rotation=45)
    plt.show()


# Funkcja do analizy playlisty i piosenki, wywoływana z głównego pliku
def analyze_song_and_save(song_name, song_text, selected_analysis):
    # Przeprowadzenie analiz
    if selected_analysis == "Sentyment":
        result = sentiment_analysis(song_text)
    elif selected_analysis == "Częstość słów":
        result = word_frequency(song_text)
        plot_word_frequency(result)  # Dodanie wykresu dla częstości słów
    elif selected_analysis == "N-gram":
        result = generate_ngrams(song_text)
        plot_ngrams(result)  # Dodanie wykresu dla n-gramów
    elif selected_analysis == "Nazwy własne":
        result = named_entity_recognition(song_text)
    elif selected_analysis == "Polaryzacja":
        result = textblob_sentiment_analysis(song_text)
    elif selected_analysis == "Składnia":
        result = syntactic_analysis(song_text)
    else:
        result = "Nie wybrano analizy"  # Domyślny wynik, jeśli żadna analiza nie została wybrana
    # Zbieranie wyników do jednej listy (aby były kompatybilne z zapisem do CSV)
    analysis_results = result

    # Przesyłanie wyników do funkcji zapisującej je w CSV
    save_to_csv(song_name, analysis_results, selected_analysis)
